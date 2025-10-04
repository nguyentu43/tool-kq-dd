import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment
from openpyxl import load_workbook

st.set_page_config(page_title="📊 Tool Điểm Danh", layout="wide")
st.title("📊 Tool Điểm Danh - Xuất Excel")

# ======================
# Hàm xử lý 1 file
# ======================
def process_file(file, buoi=None):
    df = pd.read_excel(file, skiprows=5)
    df = df.iloc[:, :-4]   # bỏ 4 cột cuối

    rows = []
    for col in df.columns[4:]:
        for lop, group in df.groupby("Lớp"):
            danh_sach = []

            # Vắng có phép (P)
            vang_p = group[group[col] == "P"]["Họ và tên"].tolist()
            danh_sach += [f"{ten} (P)" for ten in vang_p]

            # Vắng không phép (K)
            vang_k = group[group[col] == "K"]["Họ và tên"].tolist()
            danh_sach += [f"{ten} (K)" for ten in vang_k]

            so_vang = len(danh_sach)

            if so_vang == 0:
                ghi_chu = "V0"
            else:
                ghi_chu = f"V{so_vang:02d}: " + ", ".join(danh_sach)

            if buoi:  # nếu có buổi (Sáng/Chiều)
                ghi_chu = f"{buoi} {ghi_chu}"

            rows.append({
                "Lớp": lop,
                "Ngày": col,
                "Thống kê": ghi_chu
            })

    return pd.DataFrame(rows)

# ======================
# Xuất ra Excel có format
# ======================
def export_excel(pivot):
    output = BytesIO()
    pivot.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    # Đặt độ rộng cột
    ws.column_dimensions['A'].width = 12
    for col in ws.iter_cols(min_col=2, max_col=ws.max_column):
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 40

    # Căn lề + wrap text
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.freeze_panes = "B2"

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# ======================
# Tabs giao diện
# ======================
tab1, tab2 = st.tabs(["📂 Xử lý 1 file", "📂 Xử lý 2 file (Sáng + Chiều)"])

# ---- Tab 1 ----
with tab1:
    st.subheader("Xử lý 1 file Excel")
    file1 = st.file_uploader("Tải file Excel", type=["xls", "xlsx"], key="onefile")
    if file1:
        if st.button("Xử lý và Tải xuống", key="btn1"):
            df1 = process_file(file1)
            pivot1 = df1.pivot(index="Lớp", columns="Ngày", values="Thống kê").reset_index()
            result1 = export_excel(pivot1)
            st.download_button(
                label="📥 Tải file kết quả",
                data=result1,
                file_name="ketqua.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl1"
            )

# ---- Tab 2 ----
with tab2:
    st.subheader("Xử lý 2 file Excel (Sáng + Chiều)")
    col1, col2 = st.columns(2)
    with col1:
        morning_file = st.file_uploader("📂 File Buổi Sáng", type=["xls", "xlsx"], key="morning")
    with col2:
        afternoon_file = st.file_uploader("📂 File Buổi Chiều", type=["xls", "xlsx"], key="afternoon")

    if st.button("Xử lý và Tải xuống", key="btn2"):
        if not morning_file or not afternoon_file:
            st.error("⚠️ Bạn cần tải lên đủ **2 file (Sáng + Chiều)** trước khi xử lý.")
        else:
            df_morning = process_file(morning_file, "Sáng")
            df_afternoon = process_file(afternoon_file, "Chiều")

            # Gộp lại
            combined = pd.concat([df_morning, df_afternoon])
            merged = combined.groupby(["Lớp", "Ngày"])["Thống kê"].apply(lambda x: "\n".join(x)).reset_index()

            # Pivot
            pivot2 = merged.pivot(index="Lớp", columns="Ngày", values="Thống kê").reset_index()

            result2 = export_excel(pivot2)
            st.download_button(
                label="📥 Tải file kết quả (Sáng + Chiều)",
                data=result2,
                file_name="ketqua_sang_chieu.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl2"
            )
